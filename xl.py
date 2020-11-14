#!/usr/bin/env python
# encoding: utf-8

import os
import glob
import openpyxl
import pandas as pd
import codecs
from openpyxl.utils import get_column_letter

status = os.system('./cp_txt.sh')
print(status)

os.getcwd()
wb = openpyxl.Workbook()
wb.save('test.xlsx')
sheet = wb.active
wb.remove(sheet)
writer = pd.ExcelWriter('test.xlsx')


def read_txt():
    files = sorted(glob.glob('*.txt'))
    #files = sorted(os.listdir('.'))

    for filename in files:
        print(filename)
        # 按行读入，删除最后一行
        file_old = open(filename, 'r', encoding="utf-8")
        lines = [i for i in file_old]
        del lines[-1]
        file_old.close()
        # 再覆盖写入
        file_new = open(filename, 'w', encoding="utf-8")
        file_new .write(''.join(lines))
        file_new .close()

        data = pd.read_table(filename, header=0,
                             encoding='utf-8', sep='\t', index_col=None)
        # data = pd.read_table(filename, header=0, encoding='utf-8', sep='\t')
        sheetname = os.path.splitext(filename)[0]
        print(sheetname)
        wb.create_sheet(sheetname)
        data.to_excel(writer, sheet_name=sheetname, index=False)
        writer.save()

# wb.remove('Sheet')
# wb.save('test.xlsx')


if __name__ == '__main__':
    read_txt()
